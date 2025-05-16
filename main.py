import os
import re
import sys
from bs4 import BeautifulSoup
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QFileDialog,
    QStackedWidget, QFormLayout, QLineEdit, QScrollArea, QSpinBox, QMessageBox
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter

def extract_parts_with_images(soup, folder_path, source_file):
    parts = []
    img_tags = [tag for tag in soup.find_all('img') if tag.get('src', '').lower().endswith('.bmp')]
    for img_tag in img_tags:
        img_src = img_tag.get('src')
        image_path = os.path.join(folder_path, img_src)
        tr = img_tag.find_parent('tr')
        rows = []
        next_tr = tr.find_next_sibling('tr')
        while next_tr:
            if next_tr.find('img'):
                break
            rows.append(next_tr)
            next_tr = next_tr.find_next_sibling('tr')
        part = {
            "Source File": source_file,
            "Geo File": "",
            "Drawing Number": "",
            "Cutting Quantity": 0,
            "Piercing Time (s)": "",
            "Weight (kg)": 0.0,
            "Dimensions": "",
            "Image Path": image_path
        }
        for row in rows:
            cells = row.find_all('td')
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True)
            if "DRAWING NUMBER" in label:
                part["Drawing Number"] = value
            elif label.strip() == "NUMBER:":
                try:
                    part["Cutting Quantity"] = int(value)
                except:
                    part["Cutting Quantity"] = 0
            elif "PIERCING TIME" in label:
                part["Piercing Time (s)"] = value.split()[0]
            elif "WEIGHT" in label:
                try:
                    part["Weight (kg)"] = round(float(value.replace("kg", "").strip().replace(",", ".")), 3)
                except:
                    part["Weight (kg)"] = 0.0
            elif "GEOFILE NAME" in label:
                part["Geo File"] = value
            elif "DIMENSIONS" in label or "DIMENSION" in label:
                part["Dimensions"] = value
        parts.append(part)
    return parts

def extract_sheet_info(soup, source_file):
    text = soup.get_text(separator="\n")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if "INFORMATION ON SINGLE PART" in lines:
        lines = lines[:lines.index("INFORMATION ON SINGLE PART")]
    info = {
        "Source File": source_file,
        "Job Name": "",
        "Program Name": "",
        "Material": "",
        "Sheet Size (mm)": "",
        "Weight (kg)": "",
        "Total Cut Length (mm)": "",
        "Machining Time (h:min:s)": "",
        "Scrap (%)": ""
    }
    for i, line in enumerate(lines):
        if line.startswith("JOB NAME:"):
            info["Job Name"] = lines[i+1]
        elif line.startswith("PROGRAM NAME:"):
            info["Program Name"] = lines[i+1]
        elif line.startswith("MATERIAL (SHEET):"):
            info["Material"] = lines[i+1]
        elif line.startswith("BLANK:"):
            size_lines = lines[i+1:i+4]
            info["Sheet Size (mm)"] = " x ".join(s.replace("mm", "").strip() for s in size_lines)
        elif line.startswith("WEIGHT:"):
            try:
                weight = lines[i+1].replace("kg", "").strip()
                if float(weight) > 10:
                    info["Weight (kg)"] = weight
            except:
                pass
        elif line.startswith("TOTAL CUTTING LENGTH:"):
            info["Total Cut Length (mm)"] = lines[i+1].replace("mm", "").strip()
        elif line.startswith("MACHINING TIME:"):
            raw = lines[i+1].lower().strip()
            h = m = s = 0
            match_hms = re.match(r"(\d+):\s*(\d+):\s*(\d+)", raw)
            if match_hms:
                h, m, s = map(int, match_hms.groups())
            else:
                h_match = re.search(r"(\d+)\s*\[h", raw)
                m_match = re.search(r"(\d+)\s*\[min", raw)
                s_match = re.search(r"(\d+)\s*\[sec", raw)
                if h_match: h = int(h_match.group(1))
                if m_match: m = int(m_match.group(1))
                if s_match: s = int(s_match.group(1))
            info["Machining Time (h:min:s)"] = f"{h:02}:{m:02}:{s:02}"
        elif line.startswith("SCRAP:"):
            info["Scrap (%)"] = lines[i+1].replace("%", "").strip()
    return info

def apply_excel_formatting(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    ws.freeze_panes = "B2"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            # Skip if this cell is part of a merged range
            if any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges):
                continue
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

def process_multiple_htmls_with_sheet_input(folder_path, output_path, sheet_quantity_lookup, customer_name, project_name):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    raw_parts_by_file = {}
    sheet_infos = []

    for file in os.listdir(folder_path):
        if not file.lower().endswith(".html"):
            continue
        with open(os.path.join(folder_path, file), "r", encoding="iso-8859-1") as f:
            soup = BeautifulSoup(f, "html.parser")

        parts = extract_parts_with_images(soup, folder_path, file)
        multiplier = sheet_quantity_lookup.get(file, 1)
        for part in parts:
            part["Cutting Quantity"] *= multiplier

        sheet_info = extract_sheet_info(soup, file)
        sheet_info["Sheet Quantity"] = multiplier
        sheet_infos.append(sheet_info)

        for part in parts:
            part["Material"] = sheet_info.get("Material", "")
        raw_parts_by_file[file] = parts

    all_parts = {}
    for parts in raw_parts_by_file.values():
        for part in parts:
            key = part["Geo File"]
            if key in all_parts:
                all_parts[key]["Cutting Quantity"] += part["Cutting Quantity"]
            else:
                all_parts[key] = part

    wb = Workbook()
    ws_parts = wb.active
    ws_parts.title = "Parts"
    ws_summary = wb.create_sheet("Sheet Summary")

    headers = [
        "Part ID", "Customer Quantity", "Cutting Quantity", "Material",
        "Weight (kg)", "Total Weight (kg)", "Dimensions",
        "Surface Treatment", "Bending", "Welding", "Threading, Press Elements, Countersinks"
    ]

    # === Insert Customer and Project Info at the top of both sheets ===
    info_text = f"Customer: {customer_name} | Project: {project_name}"
    ws_parts.append([info_text])
    ws_parts.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers)+1)
    ws_parts['A1'].font = Font(bold=True)

    ws_parts.append(["Image"] + headers)

    for idx, part in enumerate(all_parts.values(), start=3):  # row index starts from 3 due to inserted row
        geo_file = part.get("Geo File", "")
        file_name = os.path.basename(geo_file)
        part_id = geo_file.split("\\")[-1].split("/")[-1].split("_")[0]
        part["Part ID"] = part_id

        geo_filename = os.path.basename(part.get("Geo File", "")).upper()
        suffix = geo_filename.split("_")[-1].replace(".GEO", "")
        match = re.search(r"(\d+)", suffix)
        part["Customer Quantity"] = int(match.group(1)) if match else 0

        try:
            part["Total Weight (kg)"] = round(
                float(part.get("Weight (kg)", 0)) * int(part.get("Cutting Quantity", 0)), 3)
        except:
            part["Total Weight (kg)"] = 0

        row = [""] + [part.get(h, "") for h in headers]
        ws_parts.append(row)

        img_path = part.get("Image Path", "")
        if os.path.isfile(img_path):
            try:
                img = PILImage.open(img_path)
                png_path = img_path.replace(".bmp", ".png")
                img.save(png_path)
                excel_img = ExcelImage(png_path)
                excel_img.width = 80
                excel_img.height = 80
                ws_parts.row_dimensions[idx].height = 65
                ws_parts.column_dimensions["A"].width = 14
                ws_parts.add_image(excel_img, f"A{idx}")
            except Exception as e:
                print(f"Image error in row {idx}: {e}")

    apply_excel_formatting(ws_parts)

    # === SHEET SUMMARY ===
    summary_headers = list(sheet_infos[0].keys())

    # Insert Customer and Project Info on top of sheet summary
    ws_summary.append([info_text])
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_headers))
    ws_summary['A1'].font = Font(bold=True)

    ws_summary.append(summary_headers)

    for row_idx, info in enumerate(sheet_infos, start=3):  # starts from row 3 now
        for col_idx, key in enumerate(summary_headers, start=1):
            value = info.get(key, "")
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            try:
                if key == "Weight (kg)":
                    cell.value = float(str(value).replace(",", "."))
                    cell.number_format = '#,##0.000'
                elif key == "Total Cut Length (mm)":
                    cell.value = int(float(str(value).replace(",", ".")))
                    cell.number_format = '#,##0'
                elif key == "Machining Time (h:min:s)":
                    h, m, s = map(int, value.split(":"))
                    cell.value = (h * 3600 + m * 60 + s) / 86400
                    cell.number_format = 'hh:mm:ss'
                elif key == "Scrap (%)":
                    cell.value = float(str(value).replace(",", ".")) / 100
                    cell.number_format = '0.00%'
                else:
                    cell.value = value
            except:
                cell.value = value

    # Add Overall Summary
    total_gross_weight = sum(float(info["Weight (kg)"]) * int(info["Sheet Quantity"]) for info in sheet_infos)
    total_scrap_weight = 0
    total_cutting_time_min = 0

    for info in sheet_infos:
        try:
            weight = float(info["Weight (kg)"].replace(",", "."))
            qty = int(info["Sheet Quantity"])
            scrap_percent = float(info.get("Scrap (%)", "0").replace(",", ".").replace("%", ""))
            total_scrap_weight += weight * qty * scrap_percent / 100

            time_str = info["Machining Time (h:min:s)"]
            h, m, s = map(int, time_str.split(":"))
            total_cutting_time_min += ((h * 60) + m + s / 60) * qty
        except Exception as e:
            print(f"Scrap/time calc error: {e}")

    total_net_weight = total_gross_weight - total_scrap_weight

    summary_data = [
        ("Overall Summary", ""),
        ("Total Gross Weight (kg)", round(total_gross_weight, 3)),
        ("Total Scrap (kg)", round(total_scrap_weight, 3)),
        ("Total Net Weight (kg)", round(total_net_weight, 3)),
        ("Total Cutting Time (min.)", round(total_cutting_time_min, 2)),
        ("", "")
    ]

    for i, (label, val) in enumerate(summary_data, start=1):
        ws_summary.insert_rows(2)
        ws_summary.cell(row=2, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=2, column=2, value=val)

    apply_excel_formatting(ws_summary)
    wb.save(output_path)
    print(f"✅ Excel saved to {output_path}")

# GUI and main function omitted due to space — it follows the exact layout from PySide6 GUI above.

class TruTopsApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TruTops Work Order Generator")
        self.resize(600, 400)

        self.folder_path = ""
        self.sheet_quantities = {}
        self.customer_name = ""
        self.project_name = ""

        self.stack = QStackedWidget()

        self.welcome_screen = self.build_welcome_screen()
        self.quantity_screen = self.build_quantity_screen()
        self.project_info_screen = self.build_project_info_screen()

        self.stack.addWidget(self.welcome_screen)
        self.stack.addWidget(self.quantity_screen)
        self.stack.addWidget(self.project_info_screen)

        layout = QVBoxLayout()
        layout.addWidget(self.stack)
        self.setLayout(layout)

    def build_welcome_screen(self):
        screen = QWidget()
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Welcome! Please select your TruTops folder."))
        btn = QPushButton("Select Folder")
        btn.clicked.connect(self.select_folder)
        layout.addWidget(btn)
        screen.setLayout(layout)
        return screen

    def build_quantity_screen(self):
        screen = QWidget()
        self.quantity_form = QFormLayout()
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setLayout(self.quantity_form)
        self.scroll.setWidget(inner)

        next_btn = QPushButton("Next")
        next_btn.clicked.connect(self.save_quantities)

        vbox = QVBoxLayout()
        vbox.addWidget(QLabel("Enter sheet quantity for each HTML file:"))
        vbox.addWidget(self.scroll)
        vbox.addWidget(next_btn)
        screen.setLayout(vbox)
        return screen

    def build_project_info_screen(self):
        screen = QWidget()
        layout = QFormLayout()
        self.customer_input = QLineEdit()
        self.project_input = QLineEdit()
        layout.addRow("Customer Name:", self.customer_input)
        layout.addRow("Project Name:", self.project_input)

        btn = QPushButton("Generate Excel")
        btn.clicked.connect(self.generate_excel)
        layout.addRow(btn)
        screen.setLayout(layout)
        return screen

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path = folder
            self.sheet_quantities.clear()
            while self.quantity_form.count():
                self.quantity_form.removeRow(0)
            html_files = [f for f in os.listdir(folder) if f.lower().endswith(".html")]
            self.quantity_inputs = {}
            for file in html_files:
                spin = QSpinBox()
                spin.setMinimum(1)
                spin.setMaximum(999)
                spin.setValue(1)
                self.quantity_inputs[file] = spin
                self.quantity_form.addRow(file, spin)
            self.stack.setCurrentIndex(1)

    def save_quantities(self):
        for file, spin in self.quantity_inputs.items():
            self.sheet_quantities[file] = spin.value()
        self.stack.setCurrentIndex(2)

    def generate_excel(self):
        import re
        from pathlib import Path

        self.customer_name = self.customer_input.text()
        self.project_name = self.project_input.text()

        if not self.customer_name or not self.project_name:
            QMessageBox.warning(self, "Input Error", "Please fill in customer and project name.")
            return

        def sanitize_filename(name):
            return re.sub(r'[\\/*?:"<>|]', "_", name)

        # Sanitize file name and set default path to Downloads
        sanitized_customer = sanitize_filename(self.customer_name)
        sanitized_project = sanitize_filename(self.project_name)
        default_filename = f"{sanitized_customer}_{sanitized_project}.xlsx"
        downloads_path = str(Path.home() / "Downloads")
        default_full_path = str(Path(downloads_path) / default_filename)

        # Show save dialog
        output_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_full_path,
            "Excel Files (*.xlsx)"
        )
        if not output_path:
            return

        try:
            process_multiple_htmls_with_sheet_input(
                folder_path=self.folder_path,
                output_path=output_path,
                sheet_quantity_lookup=self.sheet_quantities,
                customer_name=self.customer_name,
                project_name=self.project_name
            )
            QMessageBox.information(self, "Success", f"Excel saved to:\n{output_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = TruTopsApp()
    win.show()
    sys.exit(app.exec())            
